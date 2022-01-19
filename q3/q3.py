import openpyxl

class Person:
    def __init__(self, förnamn, efternamn, personnummer, född, epost = str(), mobilnummer = int()):
        self.förnamn = förnamn
        self.efternamn = efternamn
        self.personnummer = personnummer
        self.född = född
        self.epost = epost
        self.mobilnummer = mobilnummer
        self.djur = []


    def __str__(self):
        if len(self.djur) == 1:
            s = self.förnamn +' '+ self.efternamn + ' med personnummer ' + self.personnummer + ' har ' + str(self.djur[0].namn)
        elif len(self.djur) > 1:
            s = (self.förnamn + ' ' + self.efternamn + ' med personnummer ' + self.personnummer + ' har ' + str(self.djur[0].namn))# + ' och ' + str(self.djur[1]))
            for i in range(1,len(self.djur)):
                s += ' och ' + str(self.djur[i].namn)
        else:
            s = self.förnamn + ' ' + self.efternamn + ' har inget djur.'
        return(s)

    def Adoptera(self, djur):
        self.djur.append(djur)

class Husdjur:
    def __init__(self, ägare = str(), djur = str(), ras = str(), namn = str()):
        self.ägare = ägare
        self.djur = djur
        self.ras = ras
        self.namn = namn

    def __str__(self):
        s = self.djur + 'en '+ self.namn + ' av rasen ' + self.ras
        return s

def excelinläsning(filnamn):
    laddad_fil = openpyxl.load_workbook(filnamn)
    ark = laddad_fil.active
    laddad_fil.close()
    return ark

def skapa_personlista(ark):
    person_lista = []
    for row in ark.iter_rows(min_row=2):
        personnummer = str(row[2].value)
        född = personnummer[0:8]
        personnummer = personnummer[2:8] + "-" + personnummer[8:]
        person = Person(row[1].value, row[0].value, personnummer, född, epost=row[3].value, mobilnummer=row[4].value)
        person_lista.append(person)
    return person_lista

def paraihop_djur_person(personer,ark):
    for row in ark.iter_rows(min_row = 2):
        for person in personer:
            if person.personnummer == row[1].value:
                djur = Husdjur(djur=row[2].value, ras=row[3].value,namn=row[4].value)
                person.Adoptera(djur)
    return (personer)

def sortera_hund_katt(person_lista):
    hundägare = []
    kattägare = []
    for person in person_lista:
        for i in range(len(person.djur)):
            if person.djur[i].djur == 'hund':
                if person not in hundägare:
                    hundägare.append(person)
            elif person.djur[i].djur == 'katt':
                if person not in kattägare:
                    kattägare.append(person)
    return(hundägare,kattägare)

def spara_excel(person_lista,filnamn,djurtyp):
    wb = openpyxl.Workbook()
    ark = wb.active
    ark['A1'] = 'Efternamn'
    ark['B1'] = 'Förnamn'
    ark['C1'] = 'Födelsedag'
    ark['D1'] = 'Epost'
    ark['E1'] = 'Telenummer'
    ark['F1'] = 'Djurnamn'
    ark['G1'] = 'Djurras'
    stegare = 2
    for person in person_lista:
        if len(person.djur) == 1:
            ark.cell(row=stegare, column=1).value = person.efternamn
            ark.cell(row=stegare, column=2).value = person.förnamn
            ark.cell(row=stegare, column=3).value = int(person.född)
            ark.cell(row=stegare, column=4).value = person.epost
            ark.cell(row=stegare, column=5).value = person.mobilnummer
            ark.cell(row=stegare, column=6).value = person.djur[0].namn
            ark.cell(row=stegare, column=7).value = person.djur[0].ras
            stegare += 1
        elif len(person.djur) > 1:
            ark.cell(row=stegare, column=1).value = person.efternamn
            ark.cell(row=stegare, column=2).value = person.förnamn
            ark.cell(row=stegare, column=3).value = int(person.född)
            ark.cell(row=stegare, column=4).value = person.epost
            ark.cell(row=stegare, column=5).value = person.mobilnummer
            for djur in person.djur:
                if djur.djur == djurtyp:
                    ark.cell(row=stegare, column=6).value = djur.namn
                    ark.cell(row=stegare, column=7).value = djur.ras
                    stegare += 1
    ark['I2'] = 'Medianålder:'
    ark['J2'] = '=YEAR(TODAY()) - LEFT(MEDIAN(C2:C'+str(stegare-1)+'),4)'
    ark['I3'] = 'Medelålder:'
    ark['J3'] = '=YEAR(TODAY()) - LEFT(AVERAGE(C2:C'+str(stegare-1)+'),4)'
    ark['I4'] = 'Äldst:'
    ark['J4'] = '=YEAR(TODAY()) - LEFT(MIN(C2:C'+str(stegare-1)+'),4)'
    ark['I5'] = 'Yngst:'
    ark['J5'] = '=YEAR(TODAY()) - LEFT(MAX(C2:C'+str(stegare-1)+'),4)'
    wb.save(filename=filnamn)
    wb.close()

def main ():
    print('Laddar in excelfiler...')
    ark_HoK = excelinläsning('HoK.xlsx')
    ark_DVF = excelinläsning('DVF.xlsx')
    person_lista = skapa_personlista(ark_DVF)
    print('Sorterar upp och matchar ihop djuren med sina ägare...')
    person_lista = paraihop_djur_person(person_lista,ark_HoK)
    hundägare,kattägare = sortera_hund_katt(person_lista)
    print('Sparar ner resultatet i "Hundägare.xlsx" och "Kattägare.xlsx"...')
    spara_excel(hundägare,'Hundägare.xlsx','hund')
    spara_excel(kattägare,'Kattägare.xlsx','katt')
    print('Programmet är färdigt')

if __name__ == '__main__':
    main()