import openpyxl, random

def main():
    frö = 42
    random.seed(frö)
    namn = ['Alice', 'Maja', 'Elsa', 'Astrid', 'Wilma', 'Freja', 'Olivia', 'Selma', 'Alma', 'Ella', 'Lilly', 'Signe',
            'Vera', 'Ines', 'Alicia', 'Ebba', 'Agnes', 'Clara', 'Saga', 'Leah', 'Noah', 'William', 'Hugo', 'Lucas',
            'Liam', 'Oscar', 'Oliver', 'Matteo']
    yrken = ['Stenhuggare', 'Bibliotikarie', 'Polis']

    nytt_exceldokument = openpyxl.Workbook()
    ark = nytt_exceldokument.active

    ark['A1'] = 'Namn'
    ark['B1'] = 'Ålder'
    ark['C1'] = 'Yrke'

    ark['E2'] = 'Medianålder'
    ark['F2'] = '=MEDIAN(B2:B19)'

    ark['E4'] = 'Medelålder'
    ark['F4'] = '=AVERAGEA(B2:B19)'

    ark['E6'] = 'Äldst'
    ark['F6'] = '=MAX(B2:B19)'
    ark['E7'] = 'Född'
    ark['F7'] = "=YEAR(TODAY()) - MAX(B2:B19)"

    for rad in range(2, 20):
        ark.cell(row=rad, column=1).value = random.choice(namn)
        ark.cell(row=rad, column=2).value = random.randint(23, 64)
        ark.cell(row=rad, column=3).value = random.choice(yrken)

    # spara
    nytt_exceldokument.save('excel_forberedelse.xlsx')
if __name__ == '__main__':
    main()

