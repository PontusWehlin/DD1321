import openpyxl

def main():
    laddat_exceldokument = openpyxl.load_workbook('excel_forberedelse.xlsx')
    ark = laddat_exceldokument.active

    matris = []
    for row in ark.iter_rows( min_row = 2 ):
        radlista = []
        for cell in row:
            radlista.append( cell.value )
        matris.append( radlista )
    print("----------------- matris ----------------")
    for x in matris:
        if x[2] == 'Polis':
            print(x[0:3])
    print('Första personen heter', matris[0][0], 'och är', matris[0][2])

    dictionary_list = []
    for row in ark.iter_rows( min_row = 2 ):
        dictionary = {}
        key1 = ark.cell( row = 1, column = 1 ).value
        key2 = ark.cell( row = 1, column = 2 ).value
        key3 = ark.cell( row = 1, column = 3 ).value
        dictionary[key1] = row[0].value
        dictionary[key2] = row[1].value
        dictionary[key3] = row[2].value
        dictionary_list.append( dictionary )
    print("-------------- dictionary_list --------------")
    for x in dictionary_list:
        if x['Yrke'] == 'Polis':
            print(x)
    print('Första personen heter', dictionary_list[0]['Namn'], 'och är', dictionary_list[0]['Yrke'])


    class Person:
        def __init__(self, n, å, y):
            self.namn = n
            self.ålder = å
            self.yrke = y
    klass_lista = []
    for row in ark.iter_rows( min_row = 2):
        p = Person(row[0].value, row[1].value, row[2].value)
        klass_lista.append(p)

    print("--------------- klasslista ----------------")
    for x in klass_lista:
        if x.yrke == 'Polis':
            print(x.namn, x.ålder, x.yrke)

    print('Första personen heter', klass_lista[0].namn, 'och är', klass_lista[0].yrke)


if __name__ == '__main__':
    main()